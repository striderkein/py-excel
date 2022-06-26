from datetime import datetime
import os
import shutil
from typing import List

from dotenv import load_dotenv
import openpyxl

from my import util

def main():
  load_dotenv()
  myName = os.getenv('CURRENT_USER_NAME')

  dist = os.path.expanduser(os.getenv('DIST'))
  todayDate = datetime.now().strftime("%Y%m%d")
  todayDateSeparateWithSlash = datetime.now().strftime("%Y/%m/%d")
  templateFilePath = './template/YYYYMMDD_業務日報_.xlsx'
  distFilePath = os.path.join(dist, f"{ todayDate }_業務日報_.xlsx")
  targetSheet = os.getenv('TARGET_SHEET')

  # template からコピーして新規ファイル作成
  shutil.copyfile(templateFilePath, distFilePath)
  # ディレクトリを指定してワークブック（いわゆる Excel ファイル）を開く
  wb = openpyxl.load_workbook(distFilePath)
  sheet = wb[targetSheet]

  cell = util.convCell("B1")
  fromCell = util.convCell("A6")
  toCell = util.convCell("B6")
  asanaTaskNameCell = util.convCell("D6")
  detailCell = util.convCell("E6")
  remarksCell = util.convCell("F6")
  startRow = detailCell['row']

  repo = 'art-lesson'
  asanaTaskName = 'some asana task'
  # command = f'cd ~/dev/{ repo };git log --oneline --branches --reverse --date=short --author=\"$(git config user.name)\" --pretty=format:\"- [%ad] %h: %s\" --since=\"last day\"'
  command = f'cd ~/dev/{ repo };git log --oneline --branches --reverse --date=short --author=\"$(git config user.name)\" --pretty=format:\"%h|%s\" --since=\"last day\"'
  stream = os.popen(command)
  logs = stream.read().split('\n')

  # ヘッダ部分の出力
  # print(f"cell['row']:{ cell['row'] }, cell.['col']:{ cell['col'] }")
  # insert to cell: A1
  # op.1: with positional arguments
  # sheet.cell(row=1, column=2).value = todayDateSeparateWithSlash
  # op.2: without positional arguments
  # sheet.cell(1, 2).value = todayDateSeparateWithSlash
  sheet.cell(row=cell['row'], column=cell['col']).value = todayDateSeparateWithSlash
  sheet.cell(row=2, column=2).value = myName
  sheet.cell(row=fromCell['row'], column=fromCell['col']).value = "10:00"
  sheet.cell(row=toCell['row'], column=toCell['col']).value = "19:00"
  sheet.cell(row=asanaTaskNameCell['row'], column=asanaTaskNameCell['col']).value = asanaTaskName

  # テーブル部分の出力
  for i, log in enumerate(logs):
    print(f'{ log }')
    lo = log.split('|')
    sheet.cell(row=startRow + i, column=detailCell['col']).value = lo[1]
    sheet.cell(row=startRow + i, column=remarksCell['col']).value = lo[0]

  # 保存 → ブッククローズ
  wb.save(distFilePath)
  wb.close()

  print('all done!')

if __name__ == '__main__':
  main()
